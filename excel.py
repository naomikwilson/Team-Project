import os
from openpyxl import Workbook, load_workbook
import xlrd

user = "savilabermudez1"
folder_path = f"C:/Users/{user}/Documents/GitHub/Team-Project/excel_files"


# capital iq saves files as xls - must be saved as xlsx


for file_name in os.listdir(folder_path):
    if file_name.endswith(".xls"):
        # load workbook
        xls_file_path = os.path.join(folder_path, file_name)
        book = xlrd.open_workbook(xls_file_path)
        wb = load_workbook(filename=xls_file_path)
        # convert to xls -> xlsm
        xlsm_file_path = os.path.splitext(xls_file_path)[0]+"xlsm"
        wb.save(xlsm_file_path)

        # load the workbook again from new xlsm file
        wb = load_workbook(filename=xlsm_file_path)

        # add new sheet
        wb.create_sheet("CompCo and Benchmarking Analysis")

        # save the workbook with new sheet
        wb.save(os.path.join(folder_path, file_name))
